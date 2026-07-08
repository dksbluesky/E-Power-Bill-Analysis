#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
elec_dashboard_app.py
----------------------
簡易桌面小工具：按鈕同步電費 + 冷氣 Dashboard（呼叫既有的 update_dashboard.py）。
雙擊 啟動電費Dashboard工具.bat 開啟。
"""

import calendar as _cal
import os
import re
import shutil
import subprocess
import sys
import tkinter as tk
import zipfile
from copy import copy as _copy
from datetime import date
from pathlib import Path
from tkinter import messagebox, scrolledtext

REPO_ROOT      = Path(__file__).parent
SCRIPT         = REPO_ROOT / "update_dashboard.py"
DEFAULT_EXCEL  = Path(r"D:\AI application code\E-bill & Air con\E-Power\電費明細_by_claude_v2.xlsm")
RAW_SHEET      = "Raw Data"
DATA_ROW_START = 4

# Bi-monthly rate brackets  (monthly limits × 2)
NS_BRACKETS = [(240, 1.78), (660, 2.26), (1000, 3.17), (1400, 4.44),
               (2000, 5.17), (float("inf"), 6.25)]
S_BRACKETS  = [(240, 1.78), (660, 2.55), (1000, 3.80), (1400, 5.14),
               (2000, 6.44), (float("inf"), 8.86)]

_CALENDAR = _cal.Calendar(firstweekday=6)   # Sunday-first

def _extend_chart_ranges(excel_path: Path, new_raw_row: int,
                          bill: dict | None = None) -> None:
    """
    After adding a record at Raw Data row `new_raw_row`:
    1. Fix any #REF! formulas in the Data sheet (sheet3) for Data row dr
    2. Extend chart ranges (aligned: both category and value end at dr)
    3. Insert new row in Dashboard table (sheet1) and push footer down

    Data-sheet: Data row dr = Raw Data row - 2
    Chart alignment (post-fix): categories Data!$A$x:$A${dr-1} → $A${dr}
                                 values     Data!$X$x:$X${dr-1} → $X${dr}
    Dashboard table row for new record: dr + 1
    bill dict keys: period, season, deg, flow, public, total
    """
    tmp      = excel_path.with_suffix(".tmp.xlsm")
    dr       = new_raw_row - 2   # Data sheet row for this record
    dash_row = dr + 1             # Dashboard table row for this record

    # ── Data sheet formula XML for the new row ─────────────────────────
    rd = new_raw_row
    n  = dr
    new_data_row_xml = (
        f'<row r="{n}" ht="15" customHeight="1" s="69">'
        f'<c r="A{n}"><f>IF(\'Raw Data\'!A{rd}="","",IFERROR(\'Raw Data\'!A{rd},""))</f><v></v></c>'
        f'<c r="B{n}" s="66"><f>IF(OR(\'Raw Data\'!A{rd}="",NOT(ISNUMBER(\'Raw Data\'!G{rd}))),0,IFERROR(\'Raw Data\'!F{rd},0))</f><v></v></c>'
        f'<c r="C{n}" s="66"><f>IF(OR(\'Raw Data\'!A{rd}="",NOT(ISNUMBER(\'Raw Data\'!G{rd}))),0,IFERROR(\'Raw Data\'!G{rd},0))</f><v></v></c>'
        f'<c r="D{n}" s="66"><f>IF(\'Raw Data\'!A{rd}="",0,IFERROR(\'Raw Data\'!C{rd},0))</f><v></v></c>'
        f'<c r="E{n}" s="66"><f>IF(\'Raw Data\'!A{rd}="",0,IFERROR(\'Raw Data\'!B{rd},0))</f><v></v></c>'
        f'</row>'
    )

    # ── Dashboard table row XML for the new bill ───────────────────────
    new_dash_xml = ""
    if bill:
        period_s   = bill.get("period", "")
        season_s   = "夏季" if bill.get("season") == "夏季" else "非夏季"
        deg_v      = bill.get("deg", 0)
        flow_v     = bill.get("flow")
        public_v   = bill.get("public")
        total_v    = bill.get("total", 0)
        flow_str   = f"${int(round(flow_v)):,}"   if flow_v   is not None else "—"
        public_str = f"${int(round(public_v)):,}" if public_v is not None else "—"
        total_str  = f"${total_v:,}"
        pct_str    = f"{round(flow_v / total_v * 100, 1)}%" if flow_v and total_v else "—"
        new_dash_xml = (
            f'<row r="{dash_row}" ht="18.75" customHeight="1" s="69">'
            f'<c r="R{dash_row}" s="19" t="inlineStr"><is><t>{period_s}</t></is></c>'
            f'<c r="S{dash_row}" s="20" t="inlineStr"><is><t>{season_s}</t></is></c>'
            f'<c r="T{dash_row}" s="21" t="inlineStr"><is><t>{deg_v} 度</t></is></c>'
            f'<c r="U{dash_row}" s="22" t="inlineStr"><is><t>{flow_str}</t></is></c>'
            f'<c r="V{dash_row}" s="21" t="inlineStr"><is><t>{public_str}</t></is></c>'
            f'<c r="W{dash_row}" s="23" t="inlineStr"><is><t>{total_str}</t></is></c>'
            f'<c r="X{dash_row}" s="24" t="inlineStr"><is><t>{pct_str}</t></is></c>'
            f'</row>'
        )

    # ── Regex patterns ─────────────────────────────────────────────────
    broken_pat    = re.compile(rf'<row r="{n}"[^>]*>.*?</row>', re.DOTALL)
    # Aligned: current chart ranges end at dr-1 → extend to dr
    cat_pat       = re.compile(rf'(Data!\$A\$\d+:\$A\$){dr - 1}')
    val_pat       = re.compile(rf'(Data!\$[B-E]\$\d+:\$[B-E]\$){dr - 1}')
    # Footer row = first row containing the 📊 emoji (&#128202;) in sheet1
    footer_row_re = re.compile(r'<row r="(\d+)"[^>]*>(?:(?!</row>).)*&#128202;', re.DOTALL)

    try:
        fixed_data = fixed_charts = fixed_dash = False
        with zipfile.ZipFile(excel_path, "r") as zin, \
             zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)

                if item.filename == "xl/worksheets/sheet3.xml":
                    text = data.decode("utf-8")
                    m = broken_pat.search(text)
                    if m and "#REF!" in m.group(0):
                        text = text[:m.start()] + new_data_row_xml + text[m.end():]
                        fixed_data = True
                    data = text.encode("utf-8")

                elif item.filename.startswith("xl/charts/") and item.filename.endswith(".xml"):
                    text = data.decode("utf-8")
                    t2   = cat_pat.sub(lambda m: m.group(1) + str(dr), text)
                    t2   = val_pat.sub(lambda m: m.group(1) + str(dr), t2)
                    if t2 != text:
                        fixed_charts = True
                    data = t2.encode("utf-8")

                elif item.filename == "xl/worksheets/sheet1.xml" and new_dash_xml:
                    text = data.decode("utf-8")
                    fm   = footer_row_re.search(text)
                    if fm:
                        footer_n = int(fm.group(1))
                        # Collect footer row numbers (>= footer_n), renumber high→low
                        footer_rows = sorted(
                            {int(r) for r in re.findall(r'<row r="(\d+)"', text)
                             if int(r) >= footer_n},
                            reverse=True,
                        )
                        for rn in footer_rows:
                            text = re.sub(rf'<row r="{rn}"', f'<row r="{rn + 1}"', text)
                            text = re.sub(
                                rf'<c r="([A-Z]+){rn}"',
                                lambda m, _n=rn: f'<c r="{m.group(1)}{_n + 1}"',
                                text,
                            )
                        # Now insert new data row at dash_row before the shifted footer
                        prev_pat = re.compile(
                            rf'(<row r="{dash_row - 1}"[^>]*>.*?</row>)', re.DOTALL
                        )
                        text = prev_pat.sub(r'\1' + new_dash_xml, text)
                        fixed_dash = True
                    data = text.encode("utf-8")

                zout.writestr(item, data)

        shutil.move(str(tmp), str(excel_path))
        parts = []
        if fixed_data:   parts.append(f"Data 第 {dr} 列公式已修復")
        if fixed_charts: parts.append("圖表範圍已延伸")
        if fixed_dash:   parts.append("Dashboard 表格已新增一列")
        if parts:
            print("Excel 自動更新：" + "；".join(parts))
    except Exception as exc:
        if tmp.exists():
            tmp.unlink()
        print(f"Excel 自動更新失敗（可手動延伸）：{exc}")


# ── Helpers ───────────────────────────────────────────────────────────────────
def _calc_flow(deg: int, brackets: list) -> float:
    total, prev = 0.0, 0
    for limit, rate in brackets:
        if deg <= prev:
            break
        total += (min(deg, limit) - prev) * rate
        prev = limit
    return round(total, 1)


def _parse_roc(s: str) -> date:
    """'115/04/29' → date(2026, 4, 29)"""
    p = s.strip().split("/")
    return date(int(p[0]) + 1911, int(p[1]), int(p[2]))


def _summer_days(start: date, end: date) -> int:
    """Days in [start, end] that fall in June 1 – Sept 30."""
    count, yr = 0, start.year
    while yr <= end.year:
        ol_s = max(start, date(yr, 6, 1))
        ol_e = min(end,   date(yr, 9, 30))
        if ol_s <= ol_e:
            count += (ol_e - ol_s).days + 1
        yr += 1
    return count


def _get_last_period() -> str | None:
    """Read the last 帳單月份 from Raw Data sheet."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(DEFAULT_EXCEL), read_only=True,
                                    data_only=True, keep_vba=True)
        ws = wb[RAW_SHEET]
        last = None
        for row_cells in ws.iter_rows(min_row=DATA_ROW_START, max_col=1, values_only=True):
            if row_cells[0] and str(row_cells[0]).strip():
                last = str(row_cells[0]).strip()
            else:
                break
        wb.close()
        return last
    except Exception:
        return None


def _next_bimonth(period: str) -> str:
    """'115/05' → '115/07'  |  '115/11' → '116/01'"""
    yr, mo = int(period[:3]), int(period[4:6]) + 2
    if mo > 12:
        mo -= 12
        yr += 1
    return f"{yr}/{mo:02d}"


def _build_remarks(ver: str, node: float, season: str,
                   s_days: int, ns_days: int, start_s: str) -> str:
    """Generate 備註 matching the sheet's existing style."""
    if s_days > 0 and ns_days > 0:
        try:
            start = _parse_roc(start_s)
            mix = "5月非夏+6月夏季混合" if start.month < 6 else "9月夏季+10月非夏混合"
        except Exception:
            mix = f"混合(非夏{ns_days}天+夏{s_days}天)"
        node_txt = f"，節電獎勵${node}" if node > 0 else ""
        return f"E 實際：版本{ver}，{mix}{node_txt}"
    else:
        if node > 0:
            return f"E 實際：版本{ver}，節電獎勵${node}"
        prefix = "純夏季" if season == "夏季" else "純非夏季"
        return f"E 實際：版本{ver}，{prefix}"


# ── Calendar Popup ────────────────────────────────────────────────────────────
class CalendarPopup(tk.Toplevel):
    _WD    = ["日", "一", "二", "三", "四", "五", "六"]
    _WD_FG = ["#DC2626", "#374151", "#374151", "#374151", "#374151", "#374151", "#2563EB"]

    def __init__(self, parent, target_var: tk.StringVar):
        super().__init__(parent)
        self.title("選擇日期")
        self.transient(parent)
        self.resizable(False, False)
        self.configure(bg="#F0F4FA")
        self.target_var = target_var

        try:
            init = _parse_roc(target_var.get().strip())
        except Exception:
            init = date.today()
        self._year, self._month, self._sel = init.year, init.month, init

        self._build_ui()
        self._draw()

        self.update_idletasks()
        px, py = parent.winfo_rootx(), parent.winfo_rooty()
        dx = (parent.winfo_width()  - self.winfo_width())  // 2
        dy = (parent.winfo_height() - self.winfo_height()) // 2
        self.geometry(f"+{px + dx}+{py + dy}")

        self.bind("<Escape>", lambda e: self.destroy())
        self.grab_set()
        self.focus_set()

    def _build_ui(self):
        bg = "#F0F4FA"
        nav = tk.Frame(self, bg=bg)
        nav.pack(fill="x", padx=10, pady=8)
        tk.Button(nav, text="◀", command=self._prev, relief="flat", bg=bg,
                  font=("Arial", 11, "bold"), activebackground="#DBEAFE",
                  cursor="hand2").pack(side="left")
        self._nav_lbl = tk.Label(nav, text="", font=("Microsoft JhengHei", 11, "bold"),
                                  bg=bg, fg="#1E3A5F", width=13)
        self._nav_lbl.pack(side="left", expand=True)
        tk.Button(nav, text="▶", command=self._next, relief="flat", bg=bg,
                  font=("Arial", 11, "bold"), activebackground="#DBEAFE",
                  cursor="hand2").pack(side="right")

        hdr = tk.Frame(self, bg=bg)
        hdr.pack(padx=10)
        for i, (h, fg) in enumerate(zip(self._WD, self._WD_FG)):
            tk.Label(hdr, text=h, font=("Microsoft JhengHei", 9, "bold"),
                     bg=bg, fg=fg, width=4, anchor="center").grid(row=0, column=i)

        self._grid = tk.Frame(self, bg=bg)
        self._grid.pack(padx=10, pady=4)

        tk.Button(self, text="今天", font=("Microsoft JhengHei", 9),
                  relief="flat", bg="#E0E7FF", fg="#1E3A5F",
                  activebackground="#C7D2FE", cursor="hand2",
                  command=self._go_today).pack(pady=(0, 8))

    def _draw(self):
        for w in self._grid.winfo_children():
            w.destroy()
        self._nav_lbl.config(text=f"{self._year - 1911} 年 {self._month:02d} 月")
        today = date.today()
        for r, week in enumerate(_CALENDAR.monthdayscalendar(self._year, self._month)):
            for c, day in enumerate(week):
                if day == 0:
                    tk.Label(self._grid, text="", width=4, bg="#F0F4FA").grid(
                        row=r, column=c, padx=1, pady=2)
                    continue
                d = date(self._year, self._month, day)
                if d == self._sel:
                    bg, fg = "#2563EB", "white"
                elif d == today:
                    bg, fg = "#FEF9C3", "#92400E"
                elif c == 0:
                    bg, fg = "#F0F4FA", "#DC2626"
                elif c == 6:
                    bg, fg = "#F0F4FA", "#2563EB"
                else:
                    bg, fg = "#F0F4FA", "#374151"
                tk.Button(self._grid, text=str(day), width=4, font=("Consolas", 9),
                          bg=bg, fg=fg, relief="flat", activebackground="#BFDBFE",
                          cursor="hand2", command=lambda d=d: self._pick(d),
                          ).grid(row=r, column=c, padx=1, pady=2)

    def _pick(self, d: date):
        self.target_var.set(f"{d.year - 1911}/{d.month:02d}/{d.day:02d}")
        self.destroy()

    def _prev(self):
        if self._month == 1:
            self._month, self._year = 12, self._year - 1
        else:
            self._month -= 1
        self._draw()

    def _next(self):
        if self._month == 12:
            self._month, self._year = 1, self._year + 1
        else:
            self._month += 1
        self._draw()

    def _go_today(self):
        t = date.today()
        self._year, self._month, self._sel = t.year, t.month, t
        self._draw()


# ── Add Bill Dialog ───────────────────────────────────────────────────────────
class AddBillDialog(tk.Toplevel):
    def __init__(self, parent, on_saved=None):
        super().__init__(parent)
        self.title("新增電費帳單")
        self.geometry("460x570")
        self.configure(bg="#F0F4FA")
        self.resizable(False, False)
        self.grab_set()
        self.on_saved     = on_saved
        self._auto_season = "非夏季"
        self._s_days      = 0
        self._ns_days     = 0

        # StringVars
        self.period_var  = tk.StringVar()
        self.start_var   = tk.StringVar()
        self.end_var     = tk.StringVar()
        self.deg_var     = tk.StringVar()
        self.total_var   = tk.StringVar()
        self.flow_var    = tk.StringVar()
        self.node_var    = tk.StringVar(value="0")
        self.season_info = tk.StringVar(value="（請輸入計費起訖日和度數）")
        self.public_info = tk.StringVar(value="—")

        # Pre-fill next bi-monthly period
        last = _get_last_period()
        if last:
            self.period_var.set(_next_bimonth(last))

        # ── Layout ──
        tk.Label(self, text="➕ 新增電費帳單", font=("Microsoft JhengHei", 13, "bold"),
                 bg="#F0F4FA", fg="#1E3A5F").pack(pady=(14, 6))

        form = tk.Frame(self, bg="#F0F4FA")
        form.pack(padx=24, fill="x")

        def row(label, var, hint=""):
            f = tk.Frame(form, bg="#F0F4FA")
            f.pack(fill="x", pady=4)
            tk.Label(f, text=label, font=("Microsoft JhengHei", 10),
                     bg="#F0F4FA", width=10, anchor="w").pack(side="left")
            e = tk.Entry(f, textvariable=var, font=("Consolas", 10), width=14)
            e.pack(side="left")
            if hint:
                tk.Label(f, text=hint, font=("Microsoft JhengHei", 8),
                         bg="#F0F4FA", fg="#9CA3AF").pack(side="left", padx=(5, 0))
            return e

        def date_row(label, var, hint=""):
            f = tk.Frame(form, bg="#F0F4FA")
            f.pack(fill="x", pady=4)
            tk.Label(f, text=label, font=("Microsoft JhengHei", 10),
                     bg="#F0F4FA", width=10, anchor="w").pack(side="left")
            e = tk.Entry(f, textvariable=var, font=("Consolas", 10), width=12)
            e.pack(side="left")
            tk.Button(f, text="📅", font=("Segoe UI Emoji", 10),
                      relief="flat", bg="#F0F4FA", cursor="hand2",
                      activebackground="#DBEAFE",
                      command=lambda v=var: CalendarPopup(self, v),
                      ).pack(side="left", padx=(4, 0))
            if hint:
                tk.Label(f, text=hint, font=("Microsoft JhengHei", 8),
                         bg="#F0F4FA", fg="#9CA3AF").pack(side="left", padx=(4, 0))
            return e

        first = row("帳單月份",  self.period_var, "如 115/07")
        date_row("計費起日",  self.start_var,  "或手動輸入 115/04/29")
        date_row("計費訖日",  self.end_var,    "或手動輸入 115/07/02")
        row("用電度數",  self.deg_var,    "整數 (度)")
        row("總電費(元)", self.total_var,  "整數")

        tk.Label(form, text="── 自動計算結果 ──", font=("Microsoft JhengHei", 8),
                 bg="#F0F4FA", fg="#9CA3AF").pack(pady=(10, 2))

        sf = tk.Frame(form, bg="#F0F4FA")
        sf.pack(fill="x", pady=3)
        tk.Label(sf, text="季節", font=("Microsoft JhengHei", 10),
                 bg="#F0F4FA", width=10, anchor="w").pack(side="left")
        tk.Label(sf, textvariable=self.season_info, font=("Microsoft JhengHei", 9),
                 bg="#F0F4FA", fg="#059669", anchor="w",
                 wraplength=310, justify="left").pack(side="left")

        ff = tk.Frame(form, bg="#F0F4FA")
        ff.pack(fill="x", pady=4)
        tk.Label(ff, text="流動費用", font=("Microsoft JhengHei", 10),
                 bg="#F0F4FA", width=10, anchor="w").pack(side="left")
        tk.Entry(ff, textvariable=self.flow_var, font=("Consolas", 10), width=14).pack(side="left")
        tk.Label(ff, text="自動填入，可手動修正", font=("Microsoft JhengHei", 8),
                 bg="#F0F4FA", fg="#9CA3AF").pack(side="left", padx=(5, 0))

        pf = tk.Frame(form, bg="#F0F4FA")
        pf.pack(fill="x", pady=3)
        tk.Label(pf, text="公共費用", font=("Microsoft JhengHei", 10),
                 bg="#F0F4FA", width=10, anchor="w").pack(side="left")
        tk.Label(pf, textvariable=self.public_info, font=("Consolas", 10),
                 bg="#F0F4FA", fg="#1E40AF").pack(side="left")
        tk.Label(pf, text="= 總電費 − 流動費用", font=("Microsoft JhengHei", 8),
                 bg="#F0F4FA", fg="#9CA3AF").pack(side="left", padx=(5, 0))

        row("節電獎勵", self.node_var)

        btn_f = tk.Frame(self, bg="#F0F4FA")
        btn_f.pack(pady=14)
        tk.Button(btn_f, text="💾 寫入 Excel", font=("Microsoft JhengHei", 11),
                  bg="#16A34A", fg="white", padx=16, pady=7, relief="flat",
                  command=self._save).pack(side="left", padx=8)
        tk.Button(btn_f, text="取消", font=("Microsoft JhengHei", 11),
                  bg="#94A3B8", fg="white", padx=16, pady=7, relief="flat",
                  command=self.destroy).pack(side="left", padx=8)

        for v in (self.start_var, self.end_var, self.deg_var):
            v.trace_add("write", lambda *_: self._auto_calc())
        for v in (self.total_var, self.flow_var):
            v.trace_add("write", lambda *_: self._update_public())

        first.focus()

    # ── Auto-calculation ──────────────────────────────────────────────────────
    def _auto_calc(self):
        try:
            start = _parse_roc(self.start_var.get().strip())
            end   = _parse_roc(self.end_var.get().strip())
            deg   = int(self.deg_var.get().strip())
            if end < start or deg <= 0:
                return
        except Exception:
            return

        total_days = (end - start).days + 1
        s_days     = _summer_days(start, end)
        ns_days    = total_days - s_days
        self._s_days, self._ns_days = s_days, ns_days

        if s_days == 0:
            self._auto_season = "非夏季"
            flow = _calc_flow(deg, NS_BRACKETS)
            self.season_info.set(f"純非夏季（{total_days} 天）")
        elif ns_days == 0:
            self._auto_season = "夏季"
            flow = _calc_flow(deg, S_BRACKETS)
            self.season_info.set(f"純夏季（{total_days} 天）")
        else:
            s_deg   = round(deg * s_days  / total_days)
            ns_deg  = deg - s_deg
            s_flow  = _calc_flow(s_deg,  S_BRACKETS)
            ns_flow = _calc_flow(ns_deg, NS_BRACKETS)
            flow    = round(s_flow + ns_flow, 1)
            self._auto_season = "夏季" if s_days >= ns_days else "非夏季"
            self.season_info.set(
                f"混合：非夏 {ns_days}天/{ns_deg}度 (${ns_flow})"
                f" + 夏季 {s_days}天/{s_deg}度 (${s_flow})"
            )

        self.flow_var.set(str(flow))

    def _update_public(self):
        try:
            total = int(self.total_var.get().strip())
            flow  = float(self.flow_var.get().strip())
            self.public_info.set(f"{round(total - flow, 1)} 元")
        except Exception:
            self.public_info.set("—")

    # ── Save ──────────────────────────────────────────────────────────────────
    def _save(self):
        period  = self.period_var.get().strip()
        deg_s   = self.deg_var.get().strip()
        total_s = self.total_var.get().strip()

        if not period or not deg_s or not total_s:
            messagebox.showerror("缺少資料", "帳單月份、用電度數、總電費為必填", parent=self)
            return
        if not re.match(r"^\d{3}/\d{2}$", period):
            messagebox.showerror("格式錯誤", "帳單月份格式應為 民國年/月，如 115/07", parent=self)
            return
        try:
            deg   = int(deg_s)
            total = int(total_s)
        except ValueError:
            messagebox.showerror("格式錯誤", "用電度數和總電費須為整數", parent=self)
            return

        flow = public = None
        node = 0.0
        try:
            if self.flow_var.get().strip():
                flow   = float(self.flow_var.get().strip())
                public = round(total - flow, 1)
            if self.node_var.get().strip():
                node = float(self.node_var.get().strip())
        except ValueError:
            messagebox.showerror("格式錯誤", "費用欄位須為數字", parent=self)
            return

        mo_int = int(period[4:6])
        yr_int = int(period[:3])
        ver    = "B" if (yr_int > 114 or (yr_int == 114 and mo_int >= 10)) else "A"
        half   = period[:3] + ("上" if mo_int <= 6 else "下")

        try:
            import openpyxl
            wb = openpyxl.load_workbook(str(DEFAULT_EXCEL), keep_vba=True)
            ws = wb[RAW_SHEET]

            # Find last data row
            last_row = DATA_ROW_START - 1
            for row_cells in ws.iter_rows(min_row=DATA_ROW_START, max_col=1):
                val = row_cells[0].value
                if val is not None and str(val).strip():
                    last_row = row_cells[0].row
                else:
                    break

            new_row = last_row + 1

            # ── Copy cell style from the row above ──
            try:
                for col in range(1, 13):
                    src = ws.cell(row=last_row, column=col)
                    dst = ws.cell(row=new_row,  column=col)
                    dst.font          = _copy(src.font)
                    dst.fill          = _copy(src.fill)
                    dst.border        = _copy(src.border)
                    dst.alignment     = _copy(src.alignment)
                    dst.number_format = src.number_format
            except Exception:
                pass  # Cosmetic; proceed if style copy fails

            # ── Write values ──
            ws.cell(row=new_row, column=1,  value=period)          # A 帳單月份
            ws.cell(row=new_row, column=2,  value=deg)             # B 用電度數
            ws.cell(row=new_row, column=3,  value=total)           # C 總電費
            ws.cell(row=new_row, column=4,  value=self._auto_season)  # D 季節
            ws.cell(row=new_row, column=5,  value=ver)             # E 費率版本
            if flow   is not None:
                ws.cell(row=new_row, column=6,  value=flow)        # F 流動費用
            if public is not None:
                ws.cell(row=new_row, column=7,  value=public)      # G 公共費用
            if flow is not None and total > 0:
                ws.cell(row=new_row, column=8,                     # H 流動% (cell is % format → store fraction)
                        value=round(flow / total, 4))
            ws.cell(row=new_row, column=9,                         # I 備註
                    value=_build_remarks(ver, node, self._auto_season,
                                         self._s_days, self._ns_days,
                                         self.start_var.get().strip()))
            ws.cell(row=new_row, column=10, value=half)            # J 半年
            ws.cell(row=new_row, column=11, value=period[:3])      # K 年度
            ws.cell(row=new_row, column=12, value=node)            # L 節電獎勵

            wb.save(str(DEFAULT_EXCEL))
            wb.close()

            # Extend chart series ranges and Dashboard table
            _extend_chart_ranges(DEFAULT_EXCEL, new_row, bill={
                "period": period,
                "season": self._auto_season,
                "deg":    deg,
                "flow":   flow,
                "public": public,
                "total":  total,
            })

            messagebox.showinfo(
                "成功",
                f"已新增 {period}（第 {new_row} 列）\n"
                f"季節：{self._auto_season}　流動：{flow}　公共：{public}\n\n"
                "按「同步並上傳」更新 Dashboard",
                parent=self,
            )
            self.destroy()
            if self.on_saved:
                self.on_saved(period)

        except Exception as e:
            messagebox.showerror("寫入失敗", str(e), parent=self)


# ── Main App ──────────────────────────────────────────────────────────────────
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("電費用量 Dashboard 工具")
        self.geometry("640x480")
        self.configure(bg="#F0F4FA")

        tk.Label(self, text="⚡ 電費 Dashboard 工具", font=("Microsoft JhengHei", 14, "bold"),
                 bg="#F0F4FA", fg="#1E3A5F").pack(pady=(16, 4))
        tk.Label(self, text="從 Excel 同步電費 / 冷氣 Dashboard 資料",
                 font=("Microsoft JhengHei", 9), bg="#F0F4FA", fg="#6B7280").pack(pady=(0, 12))

        btn_frame = tk.Frame(self, bg="#F0F4FA")
        btn_frame.pack(pady=4)

        tk.Button(btn_frame, text="➕ 新增電費帳單", font=("Microsoft JhengHei", 11),
                  bg="#16A34A", fg="white", padx=14, pady=8, relief="flat",
                  command=self.add_bill).grid(row=0, column=0, padx=6)

        tk.Button(btn_frame, text="🔄 同步並上傳到 GitHub", font=("Microsoft JhengHei", 11),
                  bg="#2563EB", fg="white", padx=14, pady=8, relief="flat",
                  command=self.sync_and_push).grid(row=0, column=1, padx=6)

        tk.Button(btn_frame, text="🧪 只測試（不上傳）", font=("Microsoft JhengHei", 11),
                  bg="#94A3B8", fg="white", padx=14, pady=8, relief="flat",
                  command=self.dry_run).grid(row=0, column=2, padx=6)

        self.log = scrolledtext.ScrolledText(self, height=18, font=("Consolas", 9), bg="white")
        self.log.pack(fill="both", expand=True, padx=16, pady=12)

    def write_log(self, text):
        self.log.insert("end", text + "\n")
        self.log.see("end")
        self.update_idletasks()

    def add_bill(self):
        AddBillDialog(
            self,
            on_saved=lambda p: self.write_log(
                f"✅ {p} 已寫入 Excel，按「同步並上傳」更新 Dashboard"
            ),
        )

    def _run(self, extra_args):
        self.log.delete("1.0", "end")
        cmd = [sys.executable, str(SCRIPT)] + extra_args
        self.write_log(f"$ {' '.join(cmd)}\n")
        env = os.environ.copy()
        env["PYTHONIOENCODING"] = "utf-8"
        try:
            proc = subprocess.Popen(
                cmd, cwd=REPO_ROOT, stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT, env=env,
            )
            for raw in proc.stdout:
                self.write_log(raw.decode("utf-8", errors="replace").rstrip())
            proc.wait()
            if proc.returncode == 0:
                self.write_log("\n✅ 執行完成")
            else:
                self.write_log(f"\n❌ 執行失敗（exit code {proc.returncode}）")
        except Exception as e:
            self.write_log(f"發生錯誤：{e}")

    def sync_and_push(self):
        self._run([])

    def dry_run(self):
        self._run(["--dry-run"])


if __name__ == "__main__":
    App().mainloop()
