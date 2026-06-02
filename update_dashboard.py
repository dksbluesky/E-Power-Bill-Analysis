#!/usr/bin/env python3
"""
update_dashboard.py
-------------------
讀取電費 Excel (.xlsm)，更新 index.html 與 ac_dashboard.html，
然後 git commit + push 到 GitHub。

使用方式：
  python update_dashboard.py
  python update_dashboard.py --dry-run
"""

import argparse, re, subprocess, sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("❌ 缺少 openpyxl，請執行：pip install openpyxl")
    sys.exit(1)

DEFAULT_EXCEL  = r"D:\AI application code\E-bill & Air con\E-Power\電費明細_by_claude_v2.xlsm"
HTML_ELEC      = "index.html"
HTML_AC        = "ac_dashboard.html"
RAW_SHEET      = "Raw Data"
DATA_ROW_START = 4

# ── 讀取 Excel ────────────────────────────────────────────────────────────────
def read_excel(path: Path) -> list[dict]:
    print(f"📂 讀取 {path.name} ...")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_vba=True)

    if RAW_SHEET not in wb.sheetnames:
        print(f"❌ 找不到工作表「{RAW_SHEET}」"); sys.exit(1)

    ws = wb[RAW_SHEET]
    records = []

    for row in ws.iter_rows(min_row=DATA_ROW_START, values_only=True):
        period  = row[0]   # A
        deg     = row[1]   # B
        total   = row[2]   # C
        season  = row[3]   # D
        ver     = row[4]   # E (formula result)
        flow_v  = row[5]   # F 流動費用
        pub_v   = row[6]   # G 公共費用
        node    = row[11]  # L 節電獎勵

        if not period or str(period).strip() == "":
            continue
        period = str(period).strip()
        if not deg or not total:
            continue

        try:
            deg   = int(deg)
            total = int(total)
        except (ValueError, TypeError):
            continue

        # Season
        season = str(season).strip() if season else "非夏季"
        summer = "夏季" in season

        # Version
        if ver and str(ver).strip() in ("A", "B"):
            ver = str(ver).strip()
        else:
            yr = int(period[:3]); mo = int(period[4:6])
            ver = "B" if (yr > 114 or (yr == 114 and mo >= 10)) else "A"

        # Anomaly
        anom = (period == "113/09")

        # Flow fee - use actual value from sheet
        flow = None
        if flow_v is not None:
            try: flow = round(float(flow_v), 1)
            except: flow = None

        # 節電獎勵
        節電 = 0
        if node is not None:
            try: 節電 = round(float(node), 1)
            except: 節電 = 0

        # Public fee - use actual value from sheet
        public = None
        if pub_v is not None and str(pub_v) not in ("N/A (溢繳抵扣)", "N/A", ""):
            try: public = round(float(pub_v), 1)
            except: public = None

        # Half / year
        mo2  = int(period[4:6])
        half = period[:3] + ("上" if mo2 <= 6 else "下")
        yr2  = period[:3]

        records.append({
            "period": period, "deg": deg, "total": total,
            "summer": summer, "ver": ver, "anom": anom,
            "half": half, "year": yr2,
            "flow": flow, "節電": 節電, "public": public,
        })

    wb.close()
    print(f"✅ 讀取完成，共 {len(records)} 筆")
    return records

# ── 產生 JS 資料陣列 ──────────────────────────────────────────────────────────
def build_js_array(records: list[dict]) -> str:
    lines = ["const RAW=["]
    for r in records:
        flow_str   = str(r["flow"])   if r["flow"]   is not None else "null"
        public_str = str(r["public"]) if r["public"] is not None else "null"
        lines.append(
            f'  {{period:"{r["period"]}",deg:{r["deg"]},total:{r["total"]},'
            f'summer:{"true" if r["summer"] else "false"},'
            f'ver:"{r["ver"]}",anom:{"true" if r["anom"] else "false"},'
            f'half:"{r["half"]}",year:"{r["year"]}",'
            f'flow:{flow_str},節電:{r["節電"]},public:{public_str}}},'
        )
    if lines[-1].endswith(","):
        lines[-1] = lines[-1][:-1]
    lines.append("];")
    return "\n".join(lines)

# ── 更新 HTML ─────────────────────────────────────────────────────────────────
MARKER_START = "// DATA BLOCK — Claude Code 同步更新此區塊"
MARKER_FENCE = "// ═"

def update_html(html_path: Path, records: list[dict]) -> bool:
    content = html_path.read_text(encoding="utf-8")

    new_block = (
        "// ══════════════════════════════════════════════\n"
        "// DATA BLOCK — Claude Code 同步更新此區塊\n"
        "// ══════════════════════════════════════════════\n"
        + build_js_array(records) + "\n"
        "// ══════════════════════════════════════════════"
    )

    pattern = r"// ═+\n// DATA BLOCK.*?\n// ═+\n.*?// ═+"
    new_content, count = re.subn(pattern, new_block, content, flags=re.DOTALL)

    if count == 0:
        print(f"⚠️  {html_path.name}：找不到 DATA BLOCK 標記"); return False

    html_path.write_text(new_content, encoding="utf-8")
    print(f"✅ {html_path.name} 已更新（{len(records)} 筆）")
    return True

# ── 更新 ac_dashboard.html 費率 ───────────────────────────────────────────────
def update_ac_rates(html_path: Path, records: list[dict]) -> bool:
    """根據實際帳單重新計算加權均價並更新 ac_dashboard.html"""
    content = html_path.read_text(encoding="utf-8")

    # Summer weighted avg: use 114/09 (pure summer version A)
    summer_bills = [r for r in records if r["summer"] and not r["anom"] and r["flow"]]
    ns_bills     = [r for r in records if not r["summer"] and not r["anom"]
                    and r["flow"] and r["ver"] == "B"]

    if summer_bills:
        s_total_flow = sum(r["flow"] for r in summer_bills)
        s_total_deg  = sum(r["deg"]  for r in summer_bills)
        rate_s = round(s_total_flow / s_total_deg, 2) if s_total_deg else 2.01
    else:
        rate_s = 2.01

    if ns_bills:
        ns_total_flow = sum(r["flow"] for r in ns_bills)
        ns_total_deg  = sum(r["deg"]  for r in ns_bills)
        rate_ns = round(ns_total_flow / ns_total_deg, 2) if ns_total_deg else 1.90
    else:
        rate_ns = 1.90

    new_content = re.sub(
        r'const RATE_NS = [\d.]+;.*',
        f'const RATE_NS = {rate_ns};  // non-summer weighted avg (auto-calculated)',
        content
    )
    new_content = re.sub(
        r'const RATE_S  = [\d.]+;.*',
        f'const RATE_S  = {rate_s};  // summer weighted avg (auto-calculated)',
        new_content
    )

    if new_content == content:
        print(f"ℹ️  {html_path.name}：費率無變動")
        return True

    html_path.write_text(new_content, encoding="utf-8")
    print(f"✅ {html_path.name} 費率更新：非夏季={rate_ns}，夏季={rate_s}")
    return True

# ── Git push ──────────────────────────────────────────────────────────────────
def git_push(message: str, files: list[str]):
    for f in files:
        subprocess.run(["git", "add", f], check=True)
    result = subprocess.run(["git", "commit", "-m", message],
                            capture_output=True, text=True)
    if "nothing to commit" in result.stdout + result.stderr:
        print("ℹ️  資料未變動，不需要 push"); return
    if result.returncode != 0:
        print(f"❌ commit 失敗：{result.stderr}"); sys.exit(1)
    subprocess.run(["git", "push"], check=True)
    print("🚀 已成功 push 到 GitHub！")

# ── main ──────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel",   default=DEFAULT_EXCEL)
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--message", default="")
    args = parser.parse_args()

    repo_root  = Path(__file__).parent
    excel_path = Path(args.excel) if Path(args.excel).is_absolute() else repo_root / args.excel
    elec_html  = repo_root / HTML_ELEC
    ac_html    = repo_root / HTML_AC

    if not excel_path.exists():
        print(f"❌ 找不到 Excel：{excel_path}"); sys.exit(1)

    records = read_excel(excel_path)
    if not records:
        print("❌ 沒有讀到資料"); sys.exit(1)

    ok1 = update_html(elec_html, records)   if elec_html.exists() else True
    ok2 = update_ac_rates(ac_html, records) if ac_html.exists()   else True

    if not (ok1 and ok2): sys.exit(1)

    if args.dry_run:
        print("ℹ️  dry-run，跳過 push"); return

    latest = records[-1]["period"]
    msg    = args.message or f"update dashboard: 最新資料至 {latest}"
    git_push(msg, [HTML_ELEC, HTML_AC])
    print(f"\n✅ 完成！兩個 Dashboard 已更新至 {latest}")

if __name__ == "__main__":
    main()
